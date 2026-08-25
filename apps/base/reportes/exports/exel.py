from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

from openpyxl.utils import get_column_letter


class ExcelExporter:

    BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    HEADER_FILL = PatternFill(
        "solid",
        fgColor="D9D9D9"
    )

    # =====================================
    # EXPORT
    # =====================================

    def export(
        self,
        filename,
        title,
        tables=None,
        header_table=None,
        nested_tables=None,
        extra_content=None,
    ):

        wb = Workbook()

        ws = wb.active

        ws.title = "Reporte"

        current_row = 1

        # =====================================
        # TITULO GENERAL
        # =====================================

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=10,
        )

        cell = ws.cell(
            row=current_row,
            column=1,
            value=title,
        )

        cell.font = Font(
            bold=True,
            size=16,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws.row_dimensions[
            current_row
        ].height = 30

        current_row += 2

        # =====================================
        # CONTENIDO EXTRA
        # =====================================

        if extra_content:

            for line in extra_content:

                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=10,
                )

                cell = ws.cell(
                    row=current_row,
                    column=1,
                    value=str(line),
                )

                cell.alignment = Alignment(
                    wrap_text=True
                )

                current_row += 1

            current_row += 1

        # =====================================
        # TABLA ENCABEZADO
        # =====================================

        if header_table:

            current_row = self._render_header_table(
                ws,
                current_row,
                header_table,
            )

            current_row += 2

        # =====================================
        # TABLAS ANIDADAS
        # =====================================

        if nested_tables:

            for nested in nested_tables:

                current_row = (
                    self._render_nested_table(
                        ws,
                        current_row,
                        nested,
                    )
                )

                current_row += 2

        # =====================================
        # TABLAS NORMALES
        # =====================================

        if tables:

            for table in tables:

                current_row = (
                    self._render_table(
                        ws,
                        current_row,
                        table,
                    )
                )

                current_row += 2

        # =====================================
        # AJUSTAR COLUMNAS
        # =====================================

        self._autosize_columns(ws)

        output = BytesIO()

        wb.save(output)

        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response[
            "Content-Disposition"
        ] = (
            f'attachment; filename="{filename or title}-{timezone.localdate()}.xlsx"'
        )

        return response

    # =====================================
    # TABLA NORMAL
    # =====================================

    def _render_table(
        self,
        ws,
        current_row,
        table,
    ):

        columns = table["columns"]

        rows = table["rows"]

        title = table.get(
            "title",
            ""
        )

        total_columns = len(columns)

        if title:

            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=total_columns,
            )

            cell = ws.cell(
                row=current_row,
                column=1,
                value=title,
            )

            cell.font = Font(
                bold=True,
                size=13,
            )

            current_row += 1

        for col_num, column_name in enumerate(
            columns,
            start=1,
        ):

            cell = ws.cell(
                row=current_row,
                column=col_num,
                value=column_name,
            )

            self._apply_header_style(
                cell
            )

        current_row += 1

        for row in rows:

            for col_num, value in enumerate(
                row,
                start=1,
            ):

                cell = ws.cell(
                    row=current_row,
                    column=col_num,
                    value=value,
                )

                self._apply_data_style(
                    cell
                )

            ws.row_dimensions[
                current_row
            ].height = 25

            current_row += 1

        return current_row

    # =====================================
    # TABLA ENCABEZADO
    # =====================================

    def _render_header_table(
        self,
        ws,
        current_row,
        data,
    ):

        for row in data:

            total = len(row)

            for col, item in enumerate(
                row,
                start=1,
            ):

                header = ws.cell(
                    row=current_row,
                    column=col,
                    value=item[0],
                )

                self._apply_header_style(
                    header
                )

            current_row += 1

            for col, item in enumerate(
                row,
                start=1,
            ):

                value = ws.cell(
                    row=current_row,
                    column=col,
                    value=item[1],
                )

                self._apply_data_style(
                    value
                )

            current_row += 1

        return current_row

    # =====================================
    # TABLAS ANIDADAS
    # =====================================

    def _render_nested_table(
        self,
        ws,
        current_row,
        nested,
    ):

        # ---------
        # PADRE
        # ---------

        parent_table = {
            "columns": nested["columns"],
            "rows": nested["rows"],
        }

        current_row = self._render_table(
            ws,
            current_row,
            parent_table,
        )

        # ---------
        # HIJA
        # ---------

        child_table = {
            "columns": nested[
                "child_columns"
            ],
            "rows": nested[
                "child_rows"
            ],
        }

        current_row = self._render_table(
            ws,
            current_row,
            child_table,
        )

        return current_row

    # =====================================
    # ESTILOS
    # =====================================

    def _apply_header_style(
        self,
        cell,
    ):

        cell.font = Font(
            bold=True
        )

        cell.fill = self.HEADER_FILL

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        cell.border = self.BORDER

    def _apply_data_style(
        self,
        cell,
    ):

        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        cell.border = self.BORDER

    # =====================================
    # AUTO SIZE
    # =====================================

    def _autosize_columns(
        self,
        ws,
    ):

        for column in ws.columns:

            max_length = 0

            letter = get_column_letter(
                column[0].column
            )

            for cell in column:

                try:

                    value = str(
                        cell.value
                    )

                    max_length = max(
                        max_length,
                        len(value),
                    )

                except Exception:
                    pass

            ws.column_dimensions[
                letter
            ].width = min(
                max(
                    max_length + 5,
                    20,
                ),
                60,
            )